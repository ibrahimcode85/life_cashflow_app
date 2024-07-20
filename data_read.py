import pandas as pd
from openpyxl import load_workbook

def get_named_range_value(ws, named_range):
    # Extract the cell reference from the named range (e.g., 'Model_Point!$C$3' -> '$C$3')
    
    cell_reference = named_range.split('!')[1]
    return ws[cell_reference].value

def extract_table_from_reference(file_path, table_reference):
    # Extract the sheet name and cell range from the cell ranges reference

    sheet_name, cell_range = table_reference.split('!')
    col_start, col_end, row_start, row_end = None, None, None, None
    
    # Extract column and row information
    col_range, row_range = cell_range.split('$')[1::2], cell_range.split('$')[2::2]
    col_start, col_end = col_range[0], col_range[1]
    row_start, row_end = int(row_range[0].replace(':', '')), int(row_range[1].replace(':', ''))
    
    # Read the specified table from the sheet
    table_df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=f"{col_start}:{col_end}", header=0, skiprows=row_start - 1, nrows=row_end - row_start)
    
    return table_df

def read_pricing_model_data(file_path):
    # Load the workbook
    wb = load_workbook(filename=file_path, data_only=True)
    
    # Access the sheets
    model_point_ws  = wb['Model_Point']
    tables_ws       = wb['Tables'     ]
    
    # Extract specific named ranges 
    #   from the 'Model_Point' tab
    age                          = wb.defined_names['Age'                       ].value
    gender                       = wb.defined_names['Gender'                    ].value
    pol_year                     = wb.defined_names['Pol_Year'                  ].value
    sum_assured                  = wb.defined_names['SumAssured'                ].value
    contribution_per_year        = wb.defined_names['Contribution_perYear'      ].value
    surplus_share_to_shf         = wb.defined_names['SurplusShare_toSHF'        ].value
    surplus_share_to_participant = wb.defined_names['SurplusShare_toParticipant'].value
    #   from the 'Table' tab
    wakalah_fmc                         = wb.defined_names['Wakalah_Thrarawat'              ].value
    expense_per_contribution_per_year   = wb.defined_names['Expense_perContribution_perYear'].value
    expense_per_fund_per_year           = wb.defined_names['Expense_perFund_perYear'        ].value
    coi_loading                         = wb.defined_names['COI_Loading'                    ].value
    
    # Get the value based on the defined name
    #   from the 'Model_Point' tab
    age_value                           = get_named_range_value(model_point_ws, age                         )
    gender_value                        = get_named_range_value(model_point_ws, gender                      )
    pol_year_value                      = get_named_range_value(model_point_ws, pol_year                    )
    sum_assured_value                   = get_named_range_value(model_point_ws, sum_assured                 )
    contribution_per_year_value         = get_named_range_value(model_point_ws, contribution_per_year       )
    surplus_share_to_shf_value          = get_named_range_value(model_point_ws, surplus_share_to_shf        )
    surplus_share_to_participant_value  = get_named_range_value(model_point_ws, surplus_share_to_participant)
    #   from the 'Table' tab
    wakalah_fmc_value = get_named_range_value(tables_ws, wakalah_fmc)
    expense_per_contribution_per_year_value = get_named_range_value(tables_ws, expense_per_contribution_per_year)
    expense_per_fund_per_year_value = get_named_range_value(tables_ws, expense_per_fund_per_year)
    coi_loading_value = get_named_range_value(tables_ws, coi_loading)

    # Extract the tables using named ranges
    table_wakalah_fee_range = wb.defined_names['Tab_WakalahFee'].value
    table_mortality_range = wb.defined_names['Tab_MortalityRates'].value
    table_risk_free_rate_range = wb.defined_names['Tab_RiskFreeRates'].value
    table_lapse_range = wb.defined_names['Tab_LapseRate'].value
    
    # Convert named ranges to DataFrames
    table_wakalah_fee = extract_table_from_reference(file_path, table_wakalah_fee_range)
    table_mortality = extract_table_from_reference(file_path, table_mortality_range)
    table_risk_free_rate = extract_table_from_reference(file_path, table_risk_free_rate_range)
    table_lapse = extract_table_from_reference(file_path, table_lapse_range)
    

    # Create a dictionary to return the data
    data = {
        'Age': age_value,
        'Gender': gender_value,
        'Pol_Year': pol_year_value,
        'SumAssured': sum_assured_value,
        'Contribution_perYear': contribution_per_year_value,
        'SurplusShare_toSHF': surplus_share_to_shf_value,
        'SurplusShare_toParticipant': surplus_share_to_participant_value,
        'Table_WakalahFee': table_wakalah_fee,
        'Table_Mortality': table_mortality,
        'Table_RiskFreeRate': table_risk_free_rate,
        'Table_Lapse': table_lapse,
        'Wakalah_FMC': wakalah_fmc_value,
        'Expense_perContribution_perYear': expense_per_contribution_per_year_value,
        'Expense_perFund_perYear': expense_per_fund_per_year_value,
        'COI_Loading': coi_loading_value
    }
    return data

# Call the function to get the extracted data
# file_path = './pricing_model.xlsx'
# pricing_model_data = read_pricing_model_data(file_path)

# # Display the extracted data
# print(pricing_model_data)
